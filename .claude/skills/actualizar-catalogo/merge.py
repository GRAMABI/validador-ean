# -*- coding: utf-8 -*-
"""
merge.py — Motor de actualización del catálogo del Validador EAN desde Excel.

Formatos: Tienda Nube y LARIX (detección automática; uno o los dos por corrida).
Reglas: las de CLAUDE.md (limpieza de EAN + dígito verificador, Title Case,
alias de marca, merge no destructivo, conflictos nunca automáticos, indent=2).

Por defecto SOLO previsualiza (no escribe). Con --write escribe catalog_data.json.
En ambos casos hace un BACKUP del catálogo actual en .backups/ (con dedup y
rotación) antes de cualquier cosa. Un Excel de formato desconocido detiene todo.

Uso:
    python merge.py <excel1.xlsx> [excel2.xlsx ...]           # preview
    python merge.py <excel1.xlsx> [...] --write               # aplica

Códigos de salida: 0 ok · 3 formato desconocido · 1 error interno
"""

import sys, os, re, json, glob, shutil, subprocess
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: falta 'openpyxl' (pip install openpyxl).", file=sys.stderr)
    sys.exit(1)

# ── Ubicación del repo / rutas ──────────────────────────────
def repo_root():
    try:
        r = subprocess.run(['git', 'rev-parse', '--show-toplevel'],
                           capture_output=True, text=True)
        if r.returncode == 0 and r.stdout.strip():
            return r.stdout.strip()
    except Exception:
        pass
    return os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))

REPO       = repo_root()
CAT_PATH   = os.path.join(REPO, 'catalog_data.json')
BACKUP_DIR = os.path.join(REPO, '.backups')
KEEP_BACKUPS = 10
MAX_DROP = 0.10  # mismo umbral que el hook

# ── Marcas conocidas (CLAUDE.md) ────────────────────────────
KNOWN_BRANDS = [
    "Lüsqtoff","Black & Decker","Rust-Oleum","Dowen Pagio","Garden Plus","Green Way",
    "Delta Plus","Generac","Ultracomb","Einhell","Truper","Yelmo","Piazza","Esab","Dogo",
    "Liquitech","Tyrolit","Gamma","Acindar","Conarco","Daewoo","Bahco","Gramabi","Genebre",
    "Daihatsu","Pretul","Decakila","Pluvius","Milwaukee","Ducasse","Merclin","Hydros",
    "Konan","Dyllu","Wanke","Bosch","Dewalt","Makita","Stanley","Pintumm","Sintex","Bremen",
    "Ferrum","Crossmaster","Peugeot","Foset","Eurosound","Irimo","Trabex","Kallay","Papaiz",
    "Bronzen","Hermex","Labor","Philco","Emblem","Eslingar","Evel","Total","Crom","Kwb",
    "BTA","Iron","Ruibal","Yoly Bell","Lalelu","Yani Toys","Crown Mustang","Ken Brown",
    "Rondi","Disney","Transformers","Pin y Pon","Cry Babies","Baby Shark","Santi","Astor",
    "Crown","Pet Box","First Steps","Andicar","Rivaplast","JYF Toys","Magic Makers",
    "Caffaro","Casita de Muñecas","First Rate","John & Mary","Trucco","Xalingo",
]
# Alias -> (patrón, canónica, ambiguo?). "gramo" es unidad de peso: si es la
# ÚNICA señal, es dudoso (no se asume Gramabi por la palabra "gramo").
ALIASES = [
    (r'\blusqtoff\b', 'Lüsqtoff', False),
    (r'\blüsqtoff\b', 'Lüsqtoff', False),
    (r'\bbta\b', 'BTA', False),
    (r'\bgramo\b', 'Gramabi', True),
    (r'\brust\s*oleum\b', 'Rust-Oleum', False),
    (r'\bdeltaplus\b', 'Delta Plus', False),
]
_brand_res = [(re.compile(r'\b' + re.escape(b) + r'\b', re.I), b) for b in KNOWN_BRANDS]

# LARIX: PROVEEDOR -> marca canónica (CLAUDE.md)
LARIX_PROV = {
    'RUIBAL':'Ruibal','YANI TOYS':'Yani Toys','ANDICAR':'Andicar','RIVAPLAST SA':'Rivaplast',
    'LALELU SRL':'Lalelu','J Y F TOYS SRL':'JYF Toys','MAGIC MAKERS SA':'Magic Makers',
    'CAFFARO':'Caffaro','CASITA DE MUÑECAS':'Casita de Muñecas','YOLY BELL':'Yoly Bell',
    'FIRST RATE SA':'First Rate','JOHN & MARY':'John & Mary','TRUCCO':'Trucco','XALINGO':'Xalingo',
}

# ── Detección de marca con CONFIANZA ────────────────────────
def detect_marca(desc):
    """Devuelve (marca, confianza, motivo). confianza ∈ {'seguro','dudoso'}."""
    if not desc:
        return ('Otros', 'dudoso', 'sin descripción')
    hallados = {}  # canónica -> ambiguo?
    for pat, canon, amb in ALIASES:
        if re.search(pat, desc, re.I):
            hallados[canon] = hallados.get(canon, True) and amb
    for rgx, b in _brand_res:
        if rgx.search(desc):
            hallados[b] = False  # una marca de la lista no es ambigua por sí sola
    canons = list(hallados.keys())
    if len(canons) == 0:
        return ('Otros', 'dudoso', 'ninguna marca conocida en la descripción')
    if len(canons) == 1:
        c = canons[0]
        if hallados[c]:
            return (c, 'dudoso', 'detectada solo por un token ambiguo')
        return (c, 'seguro', '')
    return ('Otros', 'dudoso', 'varias marcas posibles: ' + ', '.join(canons))

# ── Limpieza (CLAUDE.md) ────────────────────────────────────
def clean_ean(raw):
    if raw is None:
        return ''
    s = str(raw).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return re.sub(r'\D', '', s)

def ean13_valido(ean):
    if len(ean) != 13 or not ean.isdigit():
        return None  # no aplica el chequeo (UPC-12, EAN-8, GTIN-14, vacío)
    s = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(ean[:12]))
    return (10 - s % 10) % 10 == int(ean[12])

def clean_desc(raw):
    if not raw:
        return ''
    s = re.sub(r'\s+', ' ', str(raw)).strip()
    if s.isupper() and len(s) > 3:
        s = s.title()
    return s

# ── Detección de formato ────────────────────────────────────
def _norm(h):
    if h is None: return ''
    return str(h).strip().lower().split('\n')[0]

def detectar_formato(ws_headers):
    hs = set(_norm(h) for h in ws_headers)
    if {'id_interno', 'descripcion', 'sku', 'gtin'} <= hs:
        return 'tn'
    if {'sku', 'producto', 'proveedor', 'ean'} <= hs:
        return 'larix'
    return None

def leer_excel(path):
    """Devuelve (formato, filas[]) o (None, headers_por_hoja) si no reconoce."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    diag = {}
    for ws in wb.worksheets:
        headers = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = list(row); break
        diag[ws.title] = [_norm(h) for h in (headers or [])]
        fmt = detectar_formato(headers or [])
        if fmt:
            filas = []
            idx = {_norm(h): i for i, h in enumerate(headers)}
            start = 3 if fmt == 'tn' else 2  # TN: 2 filas de header
            for row in ws.iter_rows(min_row=start, values_only=True):
                filas.append(row)
            return fmt, filas, idx
    return None, None, diag

# ── Backup con dedup + rotación ─────────────────────────────
def hacer_backup():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    prev = sorted(glob.glob(os.path.join(BACKUP_DIR, 'catalog_data.backup-*.json')))
    # dedup: si el último backup es idéntico al catálogo actual, no dupliques
    if prev:
        try:
            with open(prev[-1], 'rb') as a, open(CAT_PATH, 'rb') as b:
                if a.read() == b.read():
                    return prev[-1], False
        except Exception:
            pass
    ts = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    dst = os.path.join(BACKUP_DIR, f'catalog_data.backup-{ts}.json')
    shutil.copy2(CAT_PATH, dst)
    # rotación: conservar los últimos KEEP_BACKUPS
    allb = sorted(glob.glob(os.path.join(BACKUP_DIR, 'catalog_data.backup-*.json')))
    for old in allb[:-KEEP_BACKUPS]:
        try: os.remove(old)
        except Exception: pass
    return dst, True

# ── Validación de la SALIDA (espejo del hook del #1) ────────
def validar_salida(cat, n_antes):
    problemas = []
    n = len(cat)
    if n < 50:
        problemas.append(f'catálogo con solo {n} SKUs (parece truncado)')
    for sku, v in list(cat.items())[:100000]:
        if not isinstance(v, dict):
            problemas.append(f'{sku}: no es objeto'); continue
        e = v.get('ean', None)
        if e is None: problemas.append(f'{sku}: falta "ean"')
        elif e != '' and not str(e).isdigit(): problemas.append(f'{sku}: "ean" no numérico')
        if not str(v.get('desc', '')).strip():  problemas.append(f'{sku}: "desc" vacío')
        if not str(v.get('marca', '')).strip(): problemas.append(f'{sku}: "marca" vacío')
        if len(problemas) >= 15: break
    drop = (n_antes - n) / n_antes if n_antes else 0
    return problemas, drop

# ── MERGE ────────────────────────────────────────────────────
def procesar(excels):
    # 1) Detectar formato de cada Excel ANTES de tocar nada
    fuentes = []
    for path in excels:
        fmt, filas, idx = leer_excel(path)
        if fmt is None:
            print("FORMATO_DESCONOCIDO")
            print(f"No reconozco el Excel: {os.path.basename(path)}")
            print("Esperaba una de estas dos formas:")
            print("  • Tienda Nube: hoja con columnas id_interno | descripcion | ... | sku | gtin")
            print("  • LARIX:       hoja con columnas SKU | PRODUCTO | PROVEEDOR | ... | EAN")
            print("Lo que encontré:")
            for hoja, hs in (idx or {}).items():
                print(f"  hoja «{hoja}»: {', '.join(h for h in hs if h) or '(sin encabezados)'}")
            print("No se tocó nada.")
            sys.exit(3)
        fuentes.append((os.path.basename(path), fmt, filas, idx))

    # 2) Cargar catálogo actual + backup
    with open(CAT_PATH, encoding='utf-8') as f:
        cat = json.load(f)
    n_antes = len(cat)
    con_ean_antes = sum(1 for v in cat.values() if clean_ean(v.get('ean')))
    cat_norm = {k.strip().upper(): k for k in cat}
    backup_path, backup_nuevo = hacer_backup()

    nuevos, ean_rellenados, conflictos, dudosas, sin_desc = [], [], [], [], []
    ean_invalidos = []
    vistos = set()

    for nombre, fmt, filas, idx in fuentes:
        for row in filas:
            if not row: continue
            if fmt == 'tn':
                sku = row[idx['sku']] if idx['sku'] < len(row) else None
                desc = clean_desc(row[idx['descripcion']] if idx['descripcion'] < len(row) else '')
                ean = clean_ean(row[idx['gtin']] if idx['gtin'] < len(row) else '')
                marca, conf, motivo = detect_marca(desc)
            else:  # larix
                sku = row[idx['sku']] if idx['sku'] < len(row) else None
                desc = clean_desc(row[idx['producto']] if idx['producto'] < len(row) else '')
                ean = clean_ean(row[idx['ean']] if idx['ean'] < len(row) else '')
                prov = str(row[idx['proveedor']] if idx['proveedor'] < len(row) else '').strip().upper()
                if prov in LARIX_PROV:
                    marca, conf, motivo = LARIX_PROV[prov], 'seguro', ''
                else:
                    marca, conf, motivo = 'Otros', 'dudoso', f'proveedor no mapeado: "{prov}"'

            sku = (str(sku).strip() if sku is not None else '')
            if not sku: continue
            key = sku.upper()
            if key in vistos:  # el mismo SKU repetido entre/dentro de Excels
                continue
            vistos.add(key)

            if ean and ean13_valido(ean) is False:
                ean_invalidos.append((sku, ean, desc))

            if key in cat_norm:
                orig = cat_norm[key]
                cat_ean = clean_ean(cat[orig].get('ean'))
                if not cat_ean and ean:
                    ean_rellenados.append((orig, ean))
                elif cat_ean and ean and cat_ean != ean:
                    conflictos.append((orig, cat_ean, ean, cat[orig].get('desc', '')))
                # desc/marca existentes se CONSERVAN (no se tocan)
            else:
                if not desc:
                    sin_desc.append(sku)  # no se puede agregar sin descripción
                    continue
                nuevos.append({'sku': sku, 'ean': ean, 'desc': desc,
                               'marca': marca, 'conf': conf, 'motivo': motivo})
                if conf == 'dudoso':
                    dudosas.append((sku, marca, motivo, desc))

    # 3) Construir el catálogo propuesto (sin escribir todavía)
    propuesto = {k: dict(v) for k, v in cat.items()}
    for orig, ean in ean_rellenados:
        propuesto[orig]['ean'] = ean
    for nv in nuevos:
        propuesto[nv['sku']] = {'ean': nv['ean'], 'desc': nv['desc'], 'marca': nv['marca']}

    problemas_salida, drop = validar_salida(propuesto, n_antes)

    return {
        'fuentes': [(n, f, len(fl)) for (n, f, fl, _i) in fuentes],
        'n_antes': n_antes, 'n_despues': len(propuesto),
        'con_ean_antes': con_ean_antes,
        'con_ean_despues': con_ean_antes + len(ean_rellenados) + sum(1 for nv in nuevos if nv['ean']),
        'nuevos': nuevos, 'ean_rellenados': ean_rellenados, 'conflictos': conflictos,
        'dudosas': dudosas, 'sin_desc': sin_desc, 'ean_invalidos': ean_invalidos,
        'backup_path': backup_path, 'backup_nuevo': backup_nuevo,
        'problemas_salida': problemas_salida, 'drop': drop,
        'propuesto': propuesto,
    }

# ── REPORTE ──────────────────────────────────────────────────
def reporte(R, escribir):
    rel = lambda p: os.path.relpath(p, REPO).replace('\\', '/')
    print("=" * 66)
    print("ACTUALIZACIÓN DE CATÁLOGO — " + ("APLICANDO" if escribir else "PREVIEW (no escribe nada)"))
    print("=" * 66)
    for n, f, c in R['fuentes']:
        print(f"  Excel: {n}  [{f}]  {c} filas")
    print(f"  Backup: {rel(R['backup_path'])}" + ("" if R['backup_nuevo'] else "  (idéntico al último, no se duplicó)"))
    print(f"\n  Catálogo: {R['n_antes']} → {R['n_despues']} SKUs   "
          f"(con EAN: {R['con_ean_antes']} → {R['con_ean_despues']})")
    print(f"  + {len(R['nuevos'])} nuevos"
          + (f"  ({len(R['dudosas'])} con marca dudosa)" if R['dudosas'] else "")
          + f"\n  + {len(R['ean_rellenados'])} EAN rellenados")

    if R['conflictos']:
        print(f"\n  ⚠ CONFLICTOS DE EAN ({len(R['conflictos'])}) — NO aplicados, decidí vos:")
        for sku, viejo, nuevo, d in R['conflictos'][:40]:
            print(f"      {sku:<18} viejo {viejo:<15} nuevo {nuevo:<15} {d[:34]}")
        if len(R['conflictos']) > 40: print(f"      ... y {len(R['conflictos'])-40} más")

    if R['dudosas']:
        print(f"\n  ⚠ MARCAS A REVISAR ({len(R['dudosas'])}) — se agregan como 'Otros', corregí después:")
        for sku, marca, motivo, d in R['dudosas'][:30]:
            print(f"      {sku:<18} {motivo[:40]:<40} {d[:34]}")
        if len(R['dudosas']) > 30: print(f"      ... y {len(R['dudosas'])-30} más")

    if R['ean_invalidos']:
        print(f"\n  ⚠ EAN-13 con dígito verificador inválido ({len(R['ean_invalidos'])}):")
        for sku, ean, d in R['ean_invalidos'][:15]:
            print(f"      {sku:<18} {ean:<15} {d[:34]}")

    if R['sin_desc']:
        print(f"\n  ⚠ NO agregados (sin descripción en el Excel): {', '.join(R['sin_desc'][:20])}")

    # Chequeo espejo del hook
    print("\n  " + ("✗" if (R['problemas_salida'] or R['drop'] > MAX_DROP) else "✓") + " Validación del hook (#1):", end=" ")
    if R['drop'] > MAX_DROP:
        print(f"BLOQUEARÍA — caída de {R['drop']*100:.1f}% de SKUs (>{MAX_DROP*100:.0f}%).")
    elif R['problemas_salida']:
        print("BLOQUEARÍA —")
        for p in R['problemas_salida'][:10]:
            print("      · " + p)
    else:
        print("pasa (estructura + volumen OK).")
    print("")

# ── MAIN ─────────────────────────────────────────────────────
def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    escribir = '--write' in sys.argv
    if not args:
        print("Uso: python merge.py <excel...> [--write]", file=sys.stderr)
        sys.exit(1)

    R = procesar(args)
    reporte(R, escribir)

    if escribir:
        if R['drop'] > MAX_DROP or R['problemas_salida']:
            print("  ✗ No escribo: el resultado no pasaría el hook (ver arriba).")
            sys.exit(1)
        with open(CAT_PATH, 'w', encoding='utf-8') as f:
            json.dump(R['propuesto'], f, ensure_ascii=False, indent=2)
        print(f"  ✓ ESCRITO catalog_data.json — {R['n_despues']} SKUs.")
        print(f"    Backup del anterior: {os.path.relpath(R['backup_path'], REPO)}")
    else:
        print("  → PREVIEW. No se escribió nada. Para aplicar: agregá --write (tras tu OK).")

if __name__ == '__main__':
    try:
        main()
    except SystemExit:
        raise
    except Exception as e:
        print("merge.py error interno: " + str(e), file=sys.stderr)
        sys.exit(1)
