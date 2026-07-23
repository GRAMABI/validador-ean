# -*- coding: utf-8 -*-
"""
auditar.py — Auditor de calidad de datos del catálogo del Validador EAN.

SOLO LECTURA: lee catalog_data.json (y js/stock.js para la clasificación de
empresa) e imprime un reporte priorizado. NUNCA escribe ni corrige nada.

Separa lo ACCIONABLE del ruido conocido, y marca lo CONFIRMADO vs lo DUDOSO.
Cada ítem accionable dice DÓNDE se corrige.

Uso:  python auditar.py [export_tiendanube.xlsx]

Si se pasa un export de Tienda Nube (con la columna id_interno), el auditor
anota en cada grupo P1 cuál SKU es MÁS NUEVO (id_interno más alto = creado
después = el duplicado que suele quedar con el código viejo). Sin el export
corre igual, solo que sin ese dato.
"""

import os, re, json, sys, unicodedata, subprocess

def repo_root():
    try:
        r = subprocess.run(['git', 'rev-parse', '--show-toplevel'],
                           capture_output=True, text=True)
        if r.returncode == 0 and r.stdout.strip():
            return r.stdout.strip()
    except Exception:
        pass
    return os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))

REPO = repo_root()
CAT_PATH = os.path.join(REPO, 'catalog_data.json')
STOCK_JS = os.path.join(REPO, 'js', 'stock.js')

# Marcas de juguetería que HOY caen en GRAMABI por no estar en MARCAS_LARIX
SOSPECHOSAS_JUGUETE = ['Baby Shark', 'Transformers', 'Pin y Pon', 'Cry Babies',
                       'Pet Box', 'Rondi', 'Santi', 'Astor', 'Disney']

def norm(s):
    s = unicodedata.normalize('NFD', (s or '').lower())
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', ' ', s)).strip()

COLORS = re.compile(r'\b(negro|negra|azul|rojo|roja|verde|amarillo|amarilla|blanco|blanca|gris|plateado|plateada|celeste|naranja|rosa|violeta|dorado|dorada|turquesa|marron|fucsia|lila|cromado|cromada)\b')
MEAS = re.compile(r'(\d+(?:[.,]\d+)?)\s*(lts?|litros?|kgs?|grs?|gramos?|mm|cm|mts?|metros?|watts?|w|cc|hp|volts?|v|ah|pulgadas?|pza?s?)\b', re.I)

def medidas(desc):
    out = set()
    for m in MEAS.finditer(norm(desc)):
        num = m.group(1).replace(',', '.')
        unit = m.group(2).lower()
        unit = {'litro':'l','litros':'l','lt':'l','lts':'l','watts':'w','watt':'w',
                'metros':'m','metro':'m','mts':'m','mt':'m','gramos':'g','gramo':'g',
                'grs':'g','gr':'g','kgs':'kg','volts':'v','volt':'v',
                'pulgadas':'pulg','pulgada':'pulg'}.get(unit, unit)
        out.add(num + unit)
    return out

def core(desc):
    c = COLORS.sub(' ', norm(desc))
    c = MEAS.sub(' ', c)
    return re.sub(r'\s+', ' ', c).strip()

def check_digit_13(ean):
    if len(ean) != 13 or not ean.isdigit():
        return None
    s = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(ean[:12]))
    return (10 - s % 10) % 10 == int(ean[12])

def leer_id_interno(path):
    """{SKU_upper: id_interno} desde un export de Tienda Nube. {} si no se puede."""
    try:
        import openpyxl
    except ImportError:
        print("  (aviso: falta openpyxl, no puedo leer el id_interno del export)")
        return {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  (aviso: no pude abrir el export: {e})")
        return {}
    for ws in wb.worksheets:
        headers = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).strip().lower().split('\n')[0] if h else '' for h in row]; break
        if headers and 'id_interno' in headers and 'sku' in headers:
            i_id, i_sku = headers.index('id_interno'), headers.index('sku')
            out = {}
            for row in ws.iter_rows(min_row=3, values_only=True):
                if i_sku < len(row) and row[i_sku]:
                    sku = str(row[i_sku]).strip().upper()
                    try: out[sku] = int(row[i_id])
                    except (TypeError, ValueError): pass
            return out
    print("  (aviso: el export no tiene columnas id_interno + sku; ¿es de Tienda Nube?)")
    return {}

def marcas_larix_de_stock():
    try:
        txt = open(STOCK_JS, encoding='utf-8').read()
        m = re.search(r'MARCAS_LARIX\s*=\s*new Set\(\[(.*?)\]\)', txt, re.S)
        return set(re.findall(r"'([^']+)'", m.group(1))) if m else set()
    except Exception:
        return set()

# ── Cargar ──────────────────────────────────────────────────
cat = json.load(open(CAT_PATH, encoding='utf-8'))
N = len(cat)

# Export de TN opcional → id_interno para inferir cuál SKU es más nuevo
_export = next((a for a in sys.argv[1:] if not a.startswith('--')), None)
ID_MAP = leer_id_interno(_export) if _export else {}

# ── Agrupar por EAN ─────────────────────────────────────────
by_ean = {}
for sku, v in cat.items():
    e = (v.get('ean') or '').strip()
    if e:
        by_ean.setdefault(e, []).append(sku)
grupos = {e: ss for e, ss in by_ean.items() if len(ss) > 1}

confirmados, marca_fill, dudosos, ruido = [], [], [], 0
for e, ss in grupos.items():
    marcas = [cat[s].get('marca', '') for s in ss]
    reales = set(m for m in marcas if m and m != 'Otros')
    hay_otros = any((not m) or m == 'Otros' for m in marcas)
    if len(reales) >= 2:
        # DOS+ marcas reales distintas sobre el mismo código: grave. CONFIRMADO.
        confirmados.append((e, ss, sorted(reales)))
    elif len(reales) == 1 and hay_otros:
        # Mismo producto (variantes) donde a algunos les falta la marca: RELLENAR.
        confirmados_marca = list(reales)[0]
        marca_fill.append((e, ss, confirmados_marca))
    else:
        # Misma marca (o todas "Otros"): dudoso solo si hay señal de spec distinta.
        meds = set(frozenset(medidas(cat[s].get('desc', ''))) for s in ss)
        if len(meds) > 1:
            dudosos.append((e, ss))          # ej: 25lts vs 35lts
        else:
            ruido += 1                        # mismo producto (color / SKU duplicado)

# ── EAN-13 con dígito verificador inválido ──────────────────
dv_inval = [(s, (v.get('ean') or '').strip()) for s, v in cat.items()
            if len((v.get('ean') or '').strip()) == 13
            and (v.get('ean') or '').strip().isdigit()
            and check_digit_13((v.get('ean') or '').strip()) is False]

# ── Empresa: juguetería como GRAMABI ────────────────────────
larix = marcas_larix_de_stock()
marcas_cat = {}
for v in cat.values():
    marcas_cat[v.get('marca', 'Otros')] = marcas_cat.get(v.get('marca', 'Otros'), 0) + 1
juguete_en_gramabi = [(m, marcas_cat[m]) for m in SOSPECHOSAS_JUGUETE
                      if m in marcas_cat and m not in larix]

# ── Informativos ────────────────────────────────────────────
sin_ean = [s for s, v in cat.items() if not (v.get('ean') or '').strip()]
sin_ean_por_marca = {}
for s in sin_ean:
    mk = cat[s].get('marca', 'Otros')
    sin_ean_por_marca[mk] = sin_ean_por_marca.get(mk, 0) + 1
desc_vacias = [s for s, v in cat.items() if not str(v.get('desc', '')).strip()]
marca_vacias = [s for s, v in cat.items() if not str(v.get('marca', '')).strip()]
largos = {}
for v in cat.values():
    e = (v.get('ean') or '').strip()
    if e: largos[len(e)] = largos.get(len(e), 0) + 1
largos_raros = {k: n for k, n in largos.items() if k not in (8, 12, 13, 14)}

# ── REPORTE ─────────────────────────────────────────────────
def linea(): print('─' * 70)

def _nuevo_de(ss):
    """SKU con id_interno más alto del grupo (= más nuevo), o None si no hay datos."""
    ids = [(s, ID_MAP[s.upper()]) for s in ss if s.upper() in ID_MAP]
    if len(ids) < 2:
        return None
    return max(ids, key=lambda x: x[1])[0]

def _fila(s, nuevo):
    idtxt = f"  id:{ID_MAP[s.upper()]}" if s.upper() in ID_MAP else ("  id:—" if ID_MAP else "")
    marca = f"  ⬅ MÁS NUEVO (probable duplicado)" if s == nuevo else ""
    print(f"        {s:<18} [{cat[s].get('marca','')}] {cat[s].get('desc','')[:44]}{idtxt}{marca}")

def show_grupo(e, ss, extra=None):
    print(f"     EAN {e}" + (f"   → {extra}" if extra else ""))
    nuevo = _nuevo_de(ss)
    for s in ss:
        _fila(s, nuevo)

print('=' * 70)
print(f"AUDITORÍA DE CATÁLOGO — {N} SKUs                       (SOLO LECTURA)")
print('=' * 70)

# P1
print(f"\n🔴 P1 · Mismo código en productos de DOS MARCAS REALES distintas "
      f"({len(confirmados)})  [CONFIRMADO]")
print("     Dos fabricantes distintos no comparten un código legítimamente: o el")
print("     código está mal, o la marca está mal. → CORREGIR EN TIENDA NUBE (planilla).")
if not ID_MAP:
    print("     (pasame un export de Tienda Nube y te marco cuál SKU es MÁS NUEVO")
    print("      por id_interno — el duplicado suele ser el nuevo con el código viejo.)")
for e, ss, marcas in confirmados:
    show_grupo(e, ss, ' vs '.join(marcas))

# P1 dudosos
print(f"\n🟠 P1 (DUDOSO) · Misma marca, capacidad/medida distinta con el MISMO código "
      f"({len(dudosos)})")
print("     Capaz son productos distintos (ej. 25 vs 35 lts), capaz un SKU duplicado.")
print("     Revisá con más ojo que los confirmados. → TIENDA NUBE si son distintos.")
for e, ss in dudosos:
    show_grupo(e, ss)

# P2 marca-fill (mismo producto, hermanos sin marca)
print(f"\n🟠 P2 · Variantes del mismo producto con marca sin completar ({len(marca_fill)})")
print("     Mismo producto (mismo código, variantes de color): algunos SKUs tienen la")
print("     marca y otros quedaron en 'Otros'. No es un código mal, es completar marca.")
print("     → asignar la marca a los 'Otros' en Tienda Nube o el catálogo.")
for e, ss, marca in marca_fill:
    show_grupo(e, ss, 'completar a: ' + marca)

# P2 empresa (CÓDIGO)
print(f"\n🟠 P2 · Juguetería clasificada como GRAMABI ({len(juguete_en_gramabi)})")
print("     ⚠️  ESTA ES LA ÚNICA QUE SE CORRIGE EN CÓDIGO, NO EN UNA PLANILLA:")
print("     ⚠️  se edita  js/stock.js  (la lista MARCAS_LARIX) — es código que se")
print("     ⚠️  PUBLICA A LAS TABLETS. Pasa por el hook y por un deploy, no por TN.")
if juguete_en_gramabi:
    for m, n in juguete_en_gramabi:
        print(f"        {m:<16} {n} SKUs")

# P3 dígito verificador
print(f"\n🟡 P3 · EAN-13 con dígito verificador inválido ({len(dv_inval)})")
print("     Casi siempre un tipeo. → verificar el código físico y corregir en TN/catálogo.")
for s, e in dv_inval[:20]:
    print(f"        {s:<18} {e}  {cat[s].get('desc','')[:40]}")

# Informativos
print("\n" + "─" * 70)
print("ℹ️  INFORMATIVO (no son errores)")
print(f"   • {len(sin_ean)} SKUs sin EAN (muchos caños/perfiles sin código físico).")
top = sorted(sin_ean_por_marca.items(), key=lambda x: -x[1])[:6]
print("     Top marcas sin EAN: " + ", ".join(f"{m}({n})" for m, n in top))
print("     → los que correspondan se cargan en TN y entran con la skill /actualizar-catalogo.")
print(f"   • {ruido} grupos de EAN compartido = MISMO producto repetido (color / SKU")
print("     duplicado): NO requieren acción.")
if largos_raros:
    print(f"   • Largos de EAN atípicos: {largos_raros}")
else:
    print("   • Largos de EAN: todos válidos (8/12/13/14). Nada raro.")
print(f"   • Descripciones vacías: {len(desc_vacias)} · marcas vacías: {len(marca_vacias)}")

print("\n" + "=" * 70)
print("RESUMEN — a corregir primero:")
print(f"  1º  {len(confirmados)} en Tienda Nube — 2 marcas reales, mismo código (grave)")
print(f"  2º  {len(juguete_en_gramabi)} en CÓDIGO (js/stock.js) — toca PRODUCCIÓN, ver arriba")
print(f"  3º  {len(marca_fill)} completar marca (variantes) · {len(dv_inval)} dígito inválido")
print(f"  4º  {len(dudosos)} dudosos a revisar (muchos son packs P25/P50 = mismo producto)")
print("=" * 70)
